"""
create_excel.py  –  Generate Full_CD/input/input_parameters.xlsx.

Run once from the repo root or from Full_CD/:
    python Full_CD/create_excel.py
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

Path(__file__).parent.joinpath('input').mkdir(exist_ok=True)
OUT = Path(__file__).parent / 'input' / 'input_parameters.xlsx'

wb = openpyxl.Workbook()

hdr_font = Font(bold=True, color='FFFFFF')
hdr_fill = PatternFill('solid', fgColor='2F5496')
thin     = Side(style='thin')
border   = Border(left=thin, right=thin, top=thin, bottom=thin)

HEADERS = ['Notation', 'Parameter', 'Units', 'Value', 'Description']


def write_sheet(ws, rows, col_widths):
    for c, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = border
            cell.alignment = Alignment(horizontal='left')
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'


# ── Sheet 1: Material_Environment ────────────────────────────────────────────
ws1 = wb.active
ws1.title = 'Material_Environment'
write_sheet(ws1, [
    ('T',     'Irradiation temperature',    'K',      723.0, 'Temperature (450 C for 316 SS benchmark)'),
    ('P',     'Displacement damage rate',   'dpa/s',  1e-6,  'Production rate (dpa per second)'),
    ('rho_d', 'Network dislocation density','m^-2',   1e13,  'Point-defect sink strength'),
], [12, 32, 10, 14, 45])

# ── Sheet 2: Physical_Properties ─────────────────────────────────────────────
ws2 = wb.create_sheet('Physical_Properties')
write_sheet(ws2, [
    ('a_m',   'Lattice parameter',                    'm',        3.63e-10,'Crystal lattice constant'),
    ('E_m_v', 'Vacancy migration energy',             'eV',       1.4,     'Barrier for single-vacancy hop'),
    ('E_m_i', 'Interstitial migration energy',        'eV',       0.2,     'Barrier for single-SIA hop'),
    ('E_m_2v','Divacancy migration energy',           'eV',       0.9,     'Barrier for divacancy hop'),
    ('E_f_v', 'Vacancy formation energy',             'eV',       1.6,     'Thermal equilibrium vacancy energy'),
    ('E_f_i', 'Interstitial formation energy',        'eV',       4.08,    'Frenkel-pair interstitial energy'),
    ('E_b_2v','Divacancy binding energy',             'eV',       0.25,    'Binding energy of a divacancy'),
    ('E_b_2i','Diinterstitial binding energy',        'eV',       0.76,    'Binding energy of a diinterstitial'),
    ('nu_i',  'Interstitial attempt frequency',       's^-1',     5e12,    'Debye frequency for SIA'),
    ('nu_v',  'Vacancy attempt frequency',            's^-1',     5e13,    'Debye frequency for vacancy'),
    ('Z_v',   'Vacancy-dislocation bias factor',      '-',        1.0,     'Capture efficiency: vacancies at network dislocations'),
    ('Z_i',   'Interstitial-dislocation bias factor', '-',        1.08,    'Capture efficiency: SIAs at dislocations (bias > 1)'),
    ('g',     'Surface (void interface) energy',      'eV/m^2',   6.24e18, 'Used in thermal emission rate Gamma_cv'),
    ('C_i1',  'Interstitial combinatorial number',    '-',        84,      'Recombination collision sites around a monomer'),
], [10, 38, 12, 14, 52])

# ── Sheet 3: Model_Parameters ─────────────────────────────────────────────────
ws3 = wb.create_sheet('Model_Parameters')
write_sheet(ws3, [
    # cluster size limits
    ('Nv',                 'Maximum vacancy cluster size',               '-',    50,    'ALL modes: upper truncation of vacancy ODE system'),
    ('Ni',                 'Maximum interstitial cluster size',          '-',    100,   'ALL modes: upper truncation of interstitial ODE system'),
    # floor
    ('C_floor',            'Concentration floor (clamp)',                '-',    1e-30, 'ALL modes: prevent negative/zero concentrations'),
    # time span
    ('t_begin',            'Integration start time',                     's',    1e-6,  'ALL modes: initial time'),
    ('t_end',              'Integration end time',                       's',    1e5,   'ALL modes: final simulation time'),
    ('n_points',           'Number of output time points',               '-',    200,   'ALL modes: rows in output array'),
    ('log_time',           'Logarithmic time spacing flag',              '0/1',  1,     'ALL modes: 1=log-spaced output, 0=linear'),
    # ODE tolerances
    ('rtol',               'ODE relative tolerance',                     '-',    1e-8,  'ALL modes: relative error tolerance'),
    ('atol',               'ODE absolute tolerance',                     '-',    1e-50, 'ALL modes: absolute error tolerance'),
    # Python LSODA
    ('n_segments',         'Log-time segments (Python LSODA only)',      '-',    100,   'Python mode: number of LSODA integration segments'),
    # C++ backend
    ('backend',            'C++ integrator backend',                     '0/1',  0,     'C++ modes: 0=CVODE, 1=ARKODE'),
    ('lmm',                'CVODE linear multistep method',              '1/2',  2,     'C++ CVODE: 1=Adams, 2=BDF (stiff)'),
    ('linsol',             'Linear solver type',                         '0-2',  0,     'C++ modes: 0=Dense, 1=Banded, 2=GMRES (required for window)'),
    ('mu',                 'Upper bandwidth for banded solver',          '-',    -1,    'C++ banded: -1 auto-sets to N-1'),
    ('ml',                 'Lower bandwidth for banded solver',          '-',    -1,    'C++ banded: -1 auto-sets to N-1'),
    ('max_order',          'Maximum BDF/Adams order (0=SUNDIALS default)','- ',  0,     'C++ CVODE: 0 uses default order'),
    ('ark_table',          'ARKODE Butcher tableau ID',                  '-',    111,   'C++ ARKODE: 111=ARK548L2SA_DIRK_8_4_5'),
    # window common
    ('window_mode',        'Window solver mode',                         '0-4',  0,     'C++: 0=Full, 1=PhaseI, 2=PhaseII, 3=PhaseIII, 4=PhaseIV(OMP)'),
    ('window_check_every', 'Window expansion check frequency',           '-',    1,     'C++ window modes: check every N output steps'),
    # Phase I
    ('window_w0_v',        'Initial vacancy window size',                '-',    50,    'C++ mode>=1: initial active vacancy equations'),
    ('window_w0_i',        'Initial interstitial window size',           '-',    100,   'C++ mode>=1: initial active interstitial equations'),
    ('window_C_expand',    'Expansion concentration threshold',          '-',    1e-18, 'C++ mode>=1: expand when C[x_hi] exceeds this'),
    ('window_expand_pad',  'Additive expansion step (Phase I)',          '-',    10,    'C++ mode=1: equations added per expansion event'),
    # Phase II
    ('window_expand_factor','Geometric expansion factor (Phase II+)',    '-',    0.0,   'C++ mode>=2: x_hi *= factor (0 uses additive pad)'),
    ('window_C_contract',  'QSS contraction threshold (Phase II+)',      '-',    0.0,   'C++ mode>=2: freeze x when |dC/dt|/C < this'),
    ('window_min_active_i','Minimum active interstitial equations',      '-',    5,     'C++ mode>=2: floor on lower contraction'),
    ('window_prec',        'GMRES Jacobi preconditioner flag',           '0/1',  0,     'C++ mode>=2: 1=enable diagonal preconditioner'),
    ('window_nuc_guard',   'Nucleation guard time for contraction',      's',    0.0,   'C++ mode>=2: suppress lower contraction until t > this'),
    # Phase III
    ('window_width',       'Fixed window width (Phase III)',             '-',    500,   'C++ mode>=3: constant number of active interstitial equations'),
    ('window_t_start',     'Phase III sliding activation time',          's',    10.0,  'C++ mode>=3: window sliding starts after this time'),
    ('window_N_thresh',    'Phase III auto-activation threshold',        '-',    1000,  'C++ mode>=3: Phase III only if total N_EQ > this'),
    # Phase IV
    ('window_omp_threads', 'OpenMP thread count (Phase IV)',             '-',    0,     'C++ mode=4: 0=use OMP_NUM_THREADS env var'),
    # Dynamic Ni extension
    ('Ni_max',             'Maximum Ni for dynamic extension',           '-',    100,   'C++ modes: analytically extend beyond Ni (set > Ni to enable)'),
    ('Ni_extend_tol',      'Ni extension trigger tolerance',             '-',    0.0,   'C++ modes: 0=disabled'),
    ('Ni_extend_margin',   'Ni extension margin',                        '-',    0,     'C++ modes: extra equations added beyond trigger'),
], [22, 46, 8, 14, 55])

wb.save(OUT)
print(f'Created {OUT}')
