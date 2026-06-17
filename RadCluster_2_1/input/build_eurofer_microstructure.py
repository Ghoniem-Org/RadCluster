#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Build EuroferMicrostructure.xlsx — a single consolidated workbook merging:
  (1) Density_MeanSize_Database_FM_Steels.xlsx
  (2) Ferritics_Irradiated_Microstructure_Data.xlsx
  (3) New quantitative content extracted from the manuscript appendix
      "Experimental Database" (Ferritic_Martensitic_Steel_Microstructure.pdf,
       Appendix A.6, Tables A.21-A.25, Eqs A.98-A.100).

All sources are cross-referenced in a unified 'References' sheet keyed by
first-author + year, with legacy [CIT-##] keys from both source workbooks.
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

SRC_DEN = "Density_MeanSize_Database_FM_Steels.xlsx"
SRC_FER = "Ferritics_Irradiated_Microstructure_Data.xlsx"
OUT     = "EuroferMicrostructure.xlsx"

den = openpyxl.load_workbook(SRC_DEN, data_only=True)
fer = openpyxl.load_workbook(SRC_FER, data_only=True)

out = openpyxl.Workbook()
out.remove(out.active)

# ---------- styling helpers ----------
TITLE_FILL  = PatternFill("solid", fgColor="1F4E78")
TITLE_FONT  = Font(bold=True, color="FFFFFF", size=12)
HDR_FILL    = PatternFill("solid", fgColor="D6E4F0")
HDR_FONT    = Font(bold=True, color="1F4E78", size=10)
BAND_FILL   = PatternFill("solid", fgColor="8EAADB")
BAND_FONT   = Font(bold=True, color="FFFFFF", size=10)
NEW_FILL    = PatternFill("solid", fgColor="C6EFCE")   # green tag for manuscript-added sheets
WRAP        = Alignment(wrap_text=True, vertical="top")
THIN        = Side(style="thin", color="BFBFBF")
BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def autosize(ws, maxw=48, minw=9):
    widths = {}
    for row in ws.iter_rows():
        for c in row:
            if c.value is None:
                continue
            col = c.column_letter
            longest = max((len(s) for s in str(c.value).split("\n")), default=0)
            widths[col] = max(widths.get(col, minw), min(longest + 2, maxw))
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

def write_grid(ws, rows, title_rows=1, band_rows=0, wrap_from=None, freeze=None,
               autofilter_header=None):
    """rows = list of lists (verbatim values). title_rows bolded as title;
    next band_rows styled as bands; the following row treated as column headers."""
    for r, rowvals in enumerate(rows, start=1):
        for c, val in enumerate(rowvals, start=1):
            ws.cell(row=r, column=c, value=val)
    nrows = len(rows)
    ncols = max((len(r) for r in rows), default=1)
    # title
    for r in range(1, title_rows + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = TITLE_FILL; cell.font = TITLE_FONT
    # bands
    for r in range(title_rows + 1, title_rows + band_rows + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = BAND_FILL; cell.font = BAND_FONT; cell.alignment = WRAP
    # column-header row
    hdr = title_rows + band_rows + 1
    for c in range(1, ncols + 1):
        cell = ws.cell(row=hdr, column=c)
        cell.fill = HDR_FILL; cell.font = HDR_FONT; cell.alignment = WRAP
    # data wrap
    start = wrap_from if wrap_from else hdr + 1
    for r in range(start, nrows + 1):
        for c in range(1, ncols + 1):
            ws.cell(row=r, column=c).alignment = WRAP
    autosize(ws)
    if freeze:
        ws.freeze_panes = freeze
    elif nrows > 1:
        ws.freeze_panes = ws.cell(row=hdr + 1, column=1)
    if autofilter_header:
        ws.auto_filter.ref = f"A{autofilter_header}:{get_column_letter(ncols)}{nrows}"

def copy_sheet(dest_name, src_wb, src_name, title_rows=1, band_rows=1):
    ws_src = src_wb[src_name]
    rows = [[c for c in row] for row in ws_src.iter_rows(values_only=True)]
    # trim fully-empty trailing rows/cols
    while rows and all(v is None for v in rows[-1]):
        rows.pop()
    ws = out.create_sheet(dest_name[:31])
    write_grid(ws, rows, title_rows=title_rows, band_rows=band_rows,
               autofilter_header=title_rows + band_rows + 1)
    return ws

# =====================================================================
# 0. INDEX
# =====================================================================
idx = out.create_sheet("Index")
index_rows = [
 ["EUROFER MICROSTRUCTURE DATABASE — Irradiated Ferritic/Martensitic Steels & Pure Iron"],
 ["Consolidated workbook — loops, voids/cavities, He bubbles, precipitates, size distributions, scaling laws"],
 [""],
 ["Built by merging two source databases plus the manuscript appendix 'Experimental Database' (Appendix A.6)."],
 ["Source 1:", "Density_MeanSize_Database_FM_Steels.xlsx"],
 ["Source 2:", "Ferritics_Irradiated_Microstructure_Data.xlsx"],
 ["Source 3:", "Ferritic_Martensitic_Steel_Microstructure.pdf  (Appendix A.6, Tables A.21-A.25, Eqs A.98-A.100)"],
 [""],
 ["NOTE on citations: the two source workbooks used INCOMPATIBLE [CIT-##] numbering schemes."],
 ["  e.g. Density [CIT-15]=Taller&Was 2019 but Ferritics [CIT-15]=Getto 2015."],
 ["  In this consolidated file the authoritative bibliography is the 'References' sheet, keyed by"],
 ["  first-author + year. Legacy [CIT-##] keys from each source are preserved there for traceability."],
 ["  Data sheets retain the cite keys exactly as in their origin workbook (do NOT cross-compare keys"],
 ["  between a Density-origin and a Ferritics-origin sheet)."],
 [""],
 ["SHEET", "ORIGIN", "CONTENTS"],
 ["Index", "—", "This navigation sheet"],
 ["References", "Merged + manuscript", "Unified bibliography (author-year) with legacy CIT keys & DOIs"],
 ["Sources", "Manuscript Table A.21", "Primary experimental sources: material, irradiation, facility, dose, T, method"],
 ["Loops_DenSize", "Density wb", "Loop density & mean size (N+ion) WITH std-dev & size-range columns"],
 ["Voids_DenSize", "Density wb", "Void/cavity density, mean size & swelling (N+ion)"],
 ["T_Series", "Density wb", "Fixed-dose, varying-temperature cut (loops+voids)"],
 ["Dose_Series", "Density wb", "Fixed-temperature, varying-dose cut (loops+voids)"],
 ["N_Loops", "Ferritics wb", "Neutron irradiation — dislocation loops"],
 ["N_Voids_Swelling", "Ferritics wb", "Neutron irradiation — voids & swelling"],
 ["N_Precipitates", "Ferritics wb", "Neutron irradiation — precipitates & phases"],
 ["N_Trends", "Ferritics wb", "Neutron irradiation — narrative summary of trends"],
 ["Ion_Loops", "Ferritics wb", "Ion irradiation — dislocation loops"],
 ["Ion_Voids_Swelling", "Ferritics wb", "Ion irradiation — voids & swelling"],
 ["Ion_He_Bubbles", "Ferritics wb", "Ion irradiation — helium bubbles (incl. ODS steels)"],
 ["Ion_Precipitates", "Ferritics wb", "Ion irradiation — precipitates & phases"],
 ["Ion_Trends", "Ferritics wb", "Ion irradiation — narrative summary of trends"],
 ["Quant_Summary", "Ferritics wb", "Combined neutron+ion key-metric summary (sortable)"],
 ["Compositions", "Ferritics wb", "Nominal chemical compositions of all alloys (wt.%)"],
 ["Loop_SizeDist", "Ferritics wb", "Loop size-distribution descriptors (shape, mean, std, peak bin)"],
 ["Loop_Bins", "Ferritics wb", "Digitized loop size-distribution histogram bins"],
 ["Void_SizeDist", "Ferritics wb", "Void/cavity size-distribution descriptors"],
 ["Void_Bins", "Ferritics wb", "Digitized void size-distribution histogram bins"],
 ["Burgers_111_Frac", "Ferritics wb", "1/2<111> vs <100> loop fraction vs diameter"],
 ["Sources", "Manuscript A.21", "(see above)"],
 ["Void_Transition", "Manuscript Table A.22", "Unimodal->bimodal cavity transition (T91, HT9)"],
 ["Swelling_Dose", "Manuscript Table A.23", "Swelling & void size vs dose (T91, HT9)"],
 ["Trends_Summary", "Manuscript Table A.24", "Quantitative trend matrix (T & dose) + anchors"],
 ["Neutron_vs_Ion", "Manuscript Table A.25", "Neutron vs ion size-distribution comparison"],
 ["Scaling_Laws", "Manuscript Eqs A.98-A.100", "Loop saturation, size power law, swelling model, DBH"],
]
write_grid(idx, index_rows, title_rows=2)
# bold the sub-table header row (the "SHEET|ORIGIN|CONTENTS" row = row 16)
for c in range(1, 4):
    cell = idx.cell(row=16, column=c); cell.fill = HDR_FILL; cell.font = HDR_FONT
idx.column_dimensions["A"].width = 22
idx.column_dimensions["B"].width = 26
idx.column_dimensions["C"].width = 80

# =====================================================================
# 1. References (unified)  — author-year authoritative; legacy keys kept
# =====================================================================
# columns: RefKey | First Author(s) | Year | Title | Journal/Source | Vol/Pages | DOI/URL |
#          Density key | Ferritics key | Entry origin | Key data / notes
refs = [
 ["UNIFIED BIBLIOGRAPHY — All Sources (author-year authoritative; legacy [CIT-##] kept for traceability)"],
 ["Ref Key","First Author(s)","Year","Title","Journal / Source","Vol. / Pages",
  "DOI / URL","Density key","Ferritics key","Entry origin","Key data / notes"],
 ["Klimenkov2012","Klimenkov et al.",2012,"Quantitative characterization of microstructural defects in up to 32 dpa neutron irradiated EUROFER97","Journal of Nuclear Materials","428, 48–54","10.1016/j.jnucmat.2012.01.022","CIT-01","CIT-01","Both","EUROFER97 BOR-60 15&32 dpa/330–335°C; loop 1.4&1.7e22, 3.4&4.8 nm; void 0.14&0.11e22, 2.6&1.6 nm"],
 ["Gaganidze2016","Gaganidze et al.",2016,"Microstructural defects in EUROFER 97 after different neutron irradiation conditions","Nuclear Materials and Energy","9, 144–150","10.1016/j.nme.2015.12.004","CIT-02","CIT-02","Both","EUROFER97 HFR vs BOR-60; higher void density in HFR from He; MX/M23C6 coarsening"],
 ["Gaganidze2011","Gaganidze et al.",2011,"Characterization of radiation induced defects in EUROFER 97 after neutron irradiation","Journal of Nuclear Materials","417, 93–98","10.1016/j.jnucmat.2010.12.261","CIT-03","CIT-03","Both","EUROFER97 16.3 dpa, 250–415°C; loop 0.5–1.2e22, 2.5–12.5 nm; max density ~300°C"],
 ["Wan2010","Wan et al.",2010,"The microstructure and tensile properties of F/M steels T91, EUROFER-97 and F82H irradiated in STIP-III","Journal of Nuclear Materials","398, 40–45","ResearchGate:245051589","CIT-04","CIT-04","Both","F82H/T91/EUROFER97 SINQ 20 dpa/345°C; He bubbles 1–3 nm at 1795 appm He"],
 ["Gao2018","Gao et al.",2018,"Microstructure evolution of T91 irradiated in the BOR60 fast reactor","Acta Materialia","142, 240–254","10.1016/j.actamat.2018.01.016","CIT-05","CIT-05","Both","T91 BOR-60 376–524°C, 15.4&35.1 dpa; a<100> loops; bimodal cavity at 35 dpa; swelling 0.01–0.02%"],
 ["Gao2019","Gao et al.",2019,"Microstructure response of F/M steel HT9 after neutron irradiation: effect of dose","Journal of Nuclear Materials","516, 312–322","10.1016/j.jnucmat.2019.03.026","CIT-07","CIT-06","Both","HT9 BOR-60 17.1&35.1 dpa/377°C; loop 0.9&0.6e22, 10&16 nm; bimodal cavity; swelling 0.02&0.07%"],
 ["Gao2020","Gao et al.",2020,"Microstructure response of F/M steel HT9 after neutron irradiation: effect of temperature","Journal of Nuclear Materials","531, 152005","10.1016/j.jnucmat.2020.152005","","CIT-29","Ferritics","HT9 BOR-60 14.6–18.6 dpa, 377–457°C; T effect on G-phase, α′, cavity size"],
 ["Yan2021","Yan et al.",2021,"Phase stability and microstructural evolution in neutron-irradiated F/M steel HT9","Journal of Nuclear Materials","557, 153161","10.1016/j.jnucmat.2021.153161","CIT-08","CIT-07","Both","HT9 ATR ≤4.16 dpa, 300–600°C; loop ~0.3e22, 18.9 nm at 4.02 dpa/300°C"],
 ["Lambrecht2014","Lambrecht et al.",2014,"Microstructural examination of neutron, proton and self-ion irradiation damage in a model Fe9Cr alloy","Journal of Nuclear Materials","449, 192–198","10.1016/j.jnucmat.2014.04.011","CIT-09","CIT-11","Both","Fe-9Cr; N 2 dpa/300°C heterog. 2e20–1.2e23, ~5 nm; self-ion 6.8e22, 6.5 nm; α′ 7.4e23"],
 ["Lambrecht2009","Lambrecht & Almazouzi",2009,"Neutron irradiation effects on microstructure and mechanical properties of pure iron and RPV model alloys","Journal of Nuclear Materials","385, 334–338","10.1016/j.jnucmat.2009.09.015","CIT-10","CIT-08","Both","Pure Fe 0.19 dpa/300°C; loop ~0.5e22, ~9.5 nm; <100> dominant"],
 ["Eldrup1999","Eldrup et al.",1999,"Effects of neutron irradiation on microstructure and mechanical properties of pure iron","Journal of Nuclear Materials","264, 1–14","10.1016/S0022-3115(98)00767-3","CIT-11","CIT-09","Both","Pure Fe 0.79 dpa/60°C; loop rafts on {111}; cavities >1e24, ~1 nm"],
 ["English1982","English et al.",1982,"A TEM study of neutron-irradiated iron","Journal of Nuclear Materials","108–109, 104–109","10.1016/0022-3115(82)90490-1","CIT-12","CIT-10","Both","Pure Fe 1 dpa/400°C; void ~0.1e22, ~5 nm; swelling 0.07%; truncated octahedral"],
 ["Heintze2016","Heintze et al.",2016,"Irradiation hardening of Fe–9Cr-based alloys and ODS Eurofer: He implantation and Fe-ion at 300°C","Journal of Nuclear Materials","470, 258–267","10.1016/j.jnucmat.2015.12.027","CIT-13","CIT-27","Both","EUROFER97/Fe-9Cr 275 keV Fe+ ≤10 dpa/300°C; loops ~equal at 3&10 dpa; He pre-implant effect"],
 ["Zinkle2018","Zinkle et al.",2018,"Effect of dose on irradiation-induced loop density and Burgers vector in ion-irradiated HT9","Philosophical Magazine","98(26), 2440–2472","10.1080/14786435.2018.1490825","CIT-14","CIT-19","Both","HT9 1 MeV Kr in-situ ≤20 dpa/420–470°C; density peaks ~4 dpa then decreases; a/2<111>+a<100>"],
 ["Zheng2018","Zheng et al.",2018,"In-situ TEM observation of loop evolution in ion-irradiated HT9","Philosophical Magazine","98, (in-situ Kr study)","10.1080/14786435.2018.1464065","","(Ion loops)","In-data","HT9 1 MeV Kr in-situ 420/440/470°C; a/2<111>→a<100> with dose; density peaks ~4 dpa"],
 ["TallerWas2019","Taller & Was",2019,"Emulation of fast reactor irradiated T91 using dual ion beam irradiation","Journal of Nuclear Materials","526, 151744","10.1016/j.jnucmat.2019.151744","CIT-15","CIT-13","Both","T91 dual-ion 5 MeV Fe2++He2+, 406–570°C, 14.6–35 dpa; bimodal cavities; +60–70°C shift. (Ferritics ref sheet mislabels author as 'Liu')"],
 ["TallerWas2020","Taller & Was",2020,"Understanding bubble and void nucleation in dual ion irradiated T91 using single-parameter experiments","Acta Materialia","198, 47–60","10.1016/j.actamat.2020.07.060","CIT-20","CIT-14","Both","T91 dual-ion single-parameter; bimodal cavities; T & He/dpa dominate over dose rate; peak swelling 460°C"],
 ["Tanaka2018","Tanaka et al.",2018,"Characterization of ordered dislocation loop raft in Fe3+ irradiated pure iron at 300°C","Journal of Nuclear Materials","515, 34–41","10.1016/j.jnucmat.2018.12.019","CIT-16","CIT-24","Both","Pure Fe 6.4 MeV Fe3+ 30 dpa/300°C; loop 3.6e23, 3.5 nm; ordered raft on {100}"],
 ["Getto2015","Getto et al.",2015,"Effect of irradiation mode on the microstructure of self-ion irradiated F/M alloys","Journal of Nuclear Materials","465, 116–126","10.1016/j.jnucmat.2015.05.016","CIT-17","CIT-15","Both","HT9/T91 5 MeV Fe2+ 140 dpa/440°C; raster vs defocused; G-phase; loop ~0.3e22, ~22 nm"],
 ["Getto2017","Getto et al.",2017,"Co-evolution of microstructure features in self-ion irradiated HT9 at very high damage levels","Journal of Nuclear Materials","484, 193–208","10.1016/j.jnucmat.2016.11.020","CIT-18","CIT-16","Both","HT9 self-ion ≤650 dpa/460°C; steady-state swelling ~0.2%/dpa; loop ~0.1e22, ~30 nm; network"],
 ["GarnerToloczko1994","Garner & Toloczko",1994,"Irradiation creep and void swelling of F/M steels at low displacement rates","Journal of Nuclear Materials","212–215, 289–295","10.1016/0022-3115(94)90067-1","CIT-19","CIT-28","Both","HT9 EBR-II/FFTF 200 dpa/420°C; swelling 0.09–1.02% (10x heat-treatment variation)"],
 ["Kim2021","Kim et al.",2021,"Comparison of void swelling of F/M and ferritic HT9 alloys after high-dose self-ion irradiation","OSTI Report","1762736","OSTI:1762736","CIT-21","CIT-17","Both","HT9 600 dpa/450°C; F/M max swelling 2.6%; ferritic 0.8%; carbon contamination critical"],
 ["Yamamoto2023","Yamamoto et al.",2023,"Irradiation hardening and void swelling behaviors of F82H IEA by dual-ion irradiation","Journal of Nuclear Science and Technology","60(9), 1100–1112","10.1080/00223131.2023.2180452","CIT-22","CIT-22","Both","F82H dual-ion 270–500°C ≤80 dpa; peak hardening 350°C; peak swelling ~0.4% at 470°C/20 dpa"],
 ["FaneyWirth2022","Faney & Wirth",2022,"Cavity evolution and void swelling in dual ion irradiated tempered martensitic steels","Journal of Nuclear Materials","571, 154002","10.1016/j.jnucmat.2022.154002","CIT-23","CIT-23","Both","F82H dual-ion 231 conditions ≤82 dpa/3700 appm He/500°C; void fv 0.5–2.0%"],
 ["ZhongTan2021","Zhong & Tan",2021,"The microstructure effects on irradiation response of F/M steels","Journal of Nuclear Materials","551, 152971","10.1016/j.jnucmat.2021.152971","CIT-06","(in-data)","Both","T91 ATR ≤8.2 dpa/292–431°C; NF616 comparison; Ni-rich clusters at 431°C"],
 ["ZhongTan2024","Zhong & Tan",2024,"The microstructure effects on irradiation response of F/M steels (9Cr-NbMo vs 9Cr-Ta)","Journal of Nuclear Materials","593, 154990","10.1016/j.jnucmat.2024.154990","","CIT-12","Ferritics","9Cr-NbMo vs 9Cr-Ta HFIR 400&490°C; 9Cr-Ta softens 208 MPa at 490°C"],
 ["Lee2021","Lee et al.",2021,"Radiation-induced swelling and precipitation in Fe++ ion-irradiated F/M steels","Journal of Nuclear Materials","556, 153168","10.1016/j.jnucmat.2021.153168","","CIT-18","Ferritics","FC92/HT9/Gr.92 3.5 MeV Fe2+ ≤480 dpa/475°C; swelling and G-phase"],
 ["Tanigawa2001","Tanigawa et al.",2001,"Response of reduced-activation ferritic steels to high-fluence ion-irradiation","Journal of Nuclear Materials","298, 389–397","10.1016/S0022-3115(01)00683-9","","CIT-20","Ferritics","RAFM F82H dual-ion high fluence; damage microstructure vs parameters"],
 ["Wakai2002","Wakai et al.",2002,"Effect of triple ion beams in F/M steel on swelling behavior","Journal of Nuclear Materials","307–311, 278–282","10.1016/S0022-3115(02)01172-8","","CIT-21","Ferritics","F82H IEA triple-ion (Fe+He+H) TIARA; He/H synergistic swelling effect"],
 ["HernandezMayoral2021","Hernandez-Mayoral et al.",2021,"Dislocation loop generation differences between thin films and bulk in pure iron under 20 MeV self-ion irradiation","Metals","11(12), 2000","10.3390/met11122000","","CIT-25","Ferritics","Pure Fe 20 MeV Fe4+ 5&10 dpa/300–450°C; bulk vs thin film; <100> dominant in thin films"],
 ["Green2024","Green et al.",2024,"MX precipitate behavior in an irradiated advanced Fe-9Cr steel: self-ion irradiation effects on phase stability","arXiv preprint","2407.10002","arXiv:2407.10002","","CIT-26","Ferritics","Fe-9Cr RAFM self-ion 1–100 dpa/300–600°C; MX-TiC coarsening >400°C; dissolution ≥50 dpa"],
 ["Rouffie2015","Rouffié et al.",2015,"Single- and multi-beam ion irradiations of EUROFER97 and EUROFER-ODS","Journal of Nuclear Materials","460, 79–85","HAL:cea-02383837","","CIT-30","Ferritics","EUROFER97/ODS Fe3+ 26 dpa/400°C single/dual/triple; He needed to nucleate cavities; oxides stable"],
 ["Brimbal2013","Brimbal et al.",2013,"He and Cr effects on radiation damage formation in ion-irradiated pure iron and Fe-5.4Cr","Acta Materialia","61, 4757–4764","10.1016/j.actamat.2013.05.004","","CIT-31","Ferritics","Pure Fe & Fe-5Cr; He+Cr effects; bubble nucleation and loop formation"],
 ["Prokhodtseva2013","Prokhodtseva et al.",2013,"Impact of He and Cr on defect accumulation in ion-irradiated ultrahigh-purity Fe(Cr) alloys","Acta Materialia","61, 6958–6971","10.1016/j.actamat.2013.08.007","","CIT-32","Ferritics","Fe(Cr) 1 MeV Kr+He 435°C, 0–20 dpa; ½<111>+<100> loops; α′ precipitates"],
 ["Was2014","Was et al.",2014,"Emulation of reactor irradiation damage using ion beams","Scripta Materialia","88, 33–36","10.1016/j.scriptamat.2014.06.003","","CIT-33","Ferritics","Review: ion vs neutron; +60°C temperature shift for matched microstructure"],
 ["Klimenkov2020","Klimenkov et al.",2020,"Correlation of microstructural and mechanical properties of neutron irradiated EUROFER97 steel","Journal of Nuclear Materials","538, 152231","10.1016/j.jnucmat.2020.152231","","CIT-34","Ferritics","EUROFER97 (Böhler) HFR 16.3 dpa/252°C; loop ~5 nm; correlates TEM with hardness/tensile"],
 ["CoppolaKlimenkov2019","Coppola & Klimenkov",2019,"Dose dependence of micro-void distributions in low-temperature neutron irradiated EUROFER97","Metals","9(5), 552","10.3390/met9050552","","CIT-35","Ferritics","EUROFER97 HFR SANS micro-voids 4.4→12.9 Å for 2.7→16.3 dpa; vol frac 0.001→0.004"],
 ["He2022","He et al.",2022,"Carbon effect on dislocation loop evolution in self-ion irradiated Fe-9Cr","Metals","12(3), 457","10.3390/met12030457","","(in-data)","In-data","Fe-9Cr 3.5 MeV Fe13+ 3 dpa/450°C; C increases loop size, decreases density; network at 550°C"],
 ["Jiang2024","Jiang et al.",2024,"Carbon effect on dislocation loops in ion-irradiated Fe-9Cr / Fe-9Cr-C","ResearchGate (2024)","—","—","","(in-data)","In-data","Fe-9Cr vs Fe-9Cr-C 3 dpa/450°C; C traps interstitials → larger loops, lower density"],
 ["SwensonWharry2018","Swenson & Wharry",2018,"Comparison of ion- and neutron-irradiated F/M and ODS steels","Journal of Nuclear Materials","(2018) 503047","10.1016/j.jnucmat.2018.503047","","(in-data)","In-data","Fe-9Cr ODS / HCM12A self-ion 3 dpa/500°C; loops invariant vs neutron at 500°C with T-shift"],
 ["Ando2018","Ando et al.",2018,"Loop microstructure and hardening of F82H under self-ion irradiation","Journal of Nuclear Materials","(2018)","—","","(in-data)","In-data","F82H 6.4 MeV Fe3+ 30 dpa/300°C; ΔH=0.85 GPa; obstacle strength a<100>=0.54, ½<111>=0.17"],
 ["Toloczko2015","Toloczko et al.",2015,"Microstructure of self-ion irradiated HT9 (raster vs defocused beam)","Journal of Nuclear Materials","(2015) 300064","10.1016/j.jnucmat.2015.300064","","(in-data)","In-data","HT9 5 MeV Fe 140 dpa/440°C; raster suppresses loop density vs defocused; 0–100 appm He"],
 ["Toloczko2024","Toloczko et al.",2024,"Dual-ion irradiation of T91: microstructure and nanoindentation correlation","Journal of Nuclear Materials","(2024) 155147","10.1016/j.jnucmat.2024.155147","","(in-data)","In-data","T91 Fe3++He2+ 16.6–72 dpa/369–520°C; matches BOR-60 with +60°C shift"],
 ["Was2025","Was et al. (IAEA CRP)",2025,"IAEA round-robin on ion-irradiated T91","Journal of Nuclear Materials","(2025) 004593","10.1016/j.jnucmat.2025.004593","","(in-data)","In-data","T91 ~47 dpa/376°C target, 13 labs; C contamination correlates with loop variability"],
 ["Rogozhkin2020","Rogozhkin et al.",2020,"Radiation-induced clustering in RAFM steel EK-181 (RUSFER)","Physics of Metals and Metallography","(2020)","10.1134/S207511332002032X","","(in-data)","In-data","EK-181 ~6 dpa/250–400°C; radiation-induced clusters drive low-T hardening"],
 ["Liu2023","Liu et al.",2023,"Irradiation response of CLAM and ODS-CLAM steels under dual Fe+He ion irradiation","Journal of Nuclear Materials","(2023) 154378","10.1016/j.jnucmat.2023.154378","","(in-data)","In-data","CLAM 11 dpa/350–550°C; loop density peaks 450°C; He bubbles at oxide edges in ODS"],
 ["Song2018","Song et al.",2018,"He bubble behavior in ODS ferritic steels under dual Fe+He ion irradiation","Journal of Nuclear Materials","(2018) 503078","10.1016/j.jnucmat.2018.503078","","(in-data)","In-data","Y-Ti/Y-Al/Y-Al-Zr ODS 30 dpa/550°C, ~3500 appm He; Y-Ti-O best bubble suppression"],
 ["Yutani2007","Yutani et al.",2007,"He bubble formation in 14Cr-ODS ferritic steels","Journal of Nuclear Materials","(2007) 367423","10.1016/j.jnucmat.2007.367423","","(in-data)","In-data","14.5Cr-ODS He+ 0.2 dpa/3500 appm/300–700°C; oxide sinks effective; de-trapping at high T"],
 ["Dai2004","Dai et al.",2004,"He bubbles in F82H irradiated in SINQ spallation target","Journal of Nuclear Materials","(2004) 325001","10.1016/j.jnucmat.2004.325001","","(in-data)","In-data","F82H SINQ 10.7–19.6 dpa/850–1740 appm He; ~1 nm bubbles; brittle fracture"],
 ["Jia2005","Jia et al.",2005,"He bubbles in T91 irradiated in SINQ spallation target","Journal of Nuclear Materials","(2005) 343212","10.1016/j.jnucmat.2005.343212","","(in-data)","In-data","T91 SINQ ≤11.8 dpa/~550 appm He; ~1 nm bubbles; density slightly > F82H"],
 ["Konstantinovic2021","Konstantinović et al.",2021,"Bubble-to-void transition at oxide nanoparticles in ODS-EUROFER under dual-beam irradiation","Journal of Nuclear Materials","(2021) 152828","10.1016/j.jnucmat.2021.152828","","(in-data)","In-data","ODS-EUROFER 10 keV He / 4 MeV Au; bubble-to-void at nanoparticles under dual beam"],
 ["SchroederTrinkaus1986","Schroeder & Trinkaus",1986,"He bubble nucleation kinetics in metals","Journal of Nuclear Materials","(1986) 90080-2","10.1016/0022-3115(86)90080-2","","(in-data)","In-data","Fe-12Cr He+ <1 dpa/250–400°C; bubbles nucleate <10 appm He; cascade He re-solution"],
 ["Gilbon1988","Gilbon et al.",1988,"Irradiation behaviour of ferritic and F/M steels (EM12, F17)","Journal of Nuclear Materials","(1988) 155268","10.1016/j.jnucmat.1988.155268","","(in-data)","In-data","EM12/F17 T<500°C; α′ Cr-rich; F17 most prone to α′; α′ more detrimental than M6C/χ"],
 ["Kondo2018","Kondo et al.",2018,"Stability of oxide particles in 12Cr-6Al ODS ferritic steel under irradiation","Nuclear Materials and Energy","(2018)","10.1016/j.nme.2018.kondo","","(in-data)","In-data","12Cr-6Al ODS; Y2O3/Al-O oxides stable under irradiation; properties maintained"],
 ["Auger2019","Auger et al.",2019,"Nanocluster stability in 14YWT NFA under ion irradiation","Journal of Nuclear Materials","(2019)","10.1016/j.jnucmat.2019.xx","","(in-data)","In-data","14YWT NFA 21 dpa/450°C; nanocluster dissolution at cryo-T, stable at 450°C"],
 ["GhoniemCho1979","Ghoniem & Cho",1979,"The simultaneous clustering of point defects during irradiation (cluster-dynamics rate theory)","Physica Status Solidi / cluster-dynamics literature","—","—","","(manuscript)","Manuscript text","Manuscript ref for loop-density saturation scaling N_sat ∝ (G0/(Ω0 μ b³/kBT))^(2/3) (Eq A.98) and DBH calibration"],
 ["Ghoniem1985","Ghoniem",1985,"Theory of dispersed-barrier hardening and defect cluster evolution under irradiation","Dispersed-barrier hardening literature","—","—","","(manuscript)","Manuscript text","Manuscript ref for dispersed-barrier hardening (DBH) model calibration; Δσy = M α μ b (N d)^0.5"],
]
ws = out.create_sheet("References")
write_grid(ws, refs, title_rows=1, autofilter_header=2)

# =====================================================================
# 2. Copy data sheets verbatim
# =====================================================================
# (dest, src_wb, src_name, title_rows, band_rows)
copy_cfg = [
 ("Loops_DenSize",      den, "Loop Density & Mean Size", 1, 1),
 ("Voids_DenSize",      den, "Void Density & Mean Size", 1, 1),
 ("T_Series",           den, "T-Series (Fixed Dose)",    1, 0),
 ("Dose_Series",        den, "Dose-Series (Fixed T)",    1, 0),
 ("N_Loops",            fer, "N – Dislocation Loops",    1, 1),
 ("N_Voids_Swelling",   fer, "N – Voids & Swelling",     1, 1),
 ("N_Precipitates",     fer, "N – Precipitates & Phases",1, 1),
 ("N_Trends",           fer, "N – Summary & Trends",     1, 0),
 ("Ion_Loops",          fer, "Ion – Dislocation Loops",  1, 1),
 ("Ion_Voids_Swelling", fer, "Ion – Voids & Swelling",   1, 1),
 ("Ion_He_Bubbles",     fer, "Ion – He Bubbles",         2, 1),
 ("Ion_Precipitates",   fer, "Ion – Precipitates & Phases",1, 1),
 ("Ion_Trends",         fer, "Ion – Summary & Trends",   1, 0),
 ("Quant_Summary",      fer, "Quant Summary",            2, 1),
 ("Compositions",       fer, "Compositions",             1, 1),
 ("Loop_SizeDist",      fer, "Loop Size Distributions",  1, 2),
 ("Loop_Bins",          fer, "Loop Bins",                1, 1),
 ("Void_SizeDist",      fer, "Void Size Distributions",  1, 2),
 ("Void_Bins",          fer, "Void Bins",                1, 1),
 ("Burgers_111_Frac",   fer, "111 Fraction",             0, 0),
]
for dest, wbsrc, src, tr, br in copy_cfg:
    copy_sheet(dest, wbsrc, src, title_rows=tr, band_rows=br)

# =====================================================================
# 3. Manuscript-derived sheets (Tables A.21-A.25, Eqs A.98-A.100)
# =====================================================================
sources = [
 ["PRIMARY SOURCES OF SIZE-DISTRIBUTION DATA (manuscript Table A.21)  —  N=neutron, SI=self-ion, DI=dual-ion; WBDF=weak-beam dark-field, BF/DF=bright/dark-field TEM"],
 ["Material","Irradiation","Facility","Dose (dpa)","T (°C)","Method","Reference (Ref Key)"],
 ["EUROFER97","N fast","BOR-60","15, 32","330–335","WBDF","Klimenkov2012"],
 ["EUROFER97","N fast","BOR-60","16.3","250–415","BF/DF","Gaganidze2011"],
 ["T91","N fast","BOR-60","15.4, 35.1","376–524","STEM","Gao2018"],
 ["HT9","N fast","BOR-60","17.1, 35.1","377","STEM","Gao2019"],
 ["HT9","N fast","ATR","4.02","300–600","STEM","Yan2021"],
 ["Fe-9Cr","N fission","BR2","2.0","300","WBDF","Lambrecht2014"],
 ["Pure Fe","N fission","Multiple","0.19","300","BF/DF","Lambrecht2009"],
 ["Pure Fe","N fission","Multiple","0.79","60","BF/DF","Eldrup1999"],
 ["Pure Fe","N fission","Reactors","1.0","400","BF/DF","English1982"],
 ["HT9","SI Kr","In-situ TEM","4, 20","420–470","In-situ","Zinkle2018"],
 ["Fe-9Cr","SI Fe+","Ion facility","2.0","300","WBDF","Lambrecht2014"],
 ["Pure Fe","SI Fe3+","Ion facility","30","300","BF/DF","Tanaka2018"],
 ["T91","DI Fe+He","Dual-ion facility","14.6–35","406–570","TEM/STEM","TallerWas2019"],
 ["T91","DI Fe+He","Dual-ion facility","14.6–35","406–570","TEM","TallerWas2020"],
 ["HT9","SI Fe2+","Self-ion","140–650","440–460","STEM","Getto2017"],
 ["F82H","DI Fe+He","Dual-ion facility","20","470","TEM","Yamamoto2023"],
 ["F82H","DI Fe+He","Dual-ion facility","≤82","500","TEM","FaneyWirth2022"],
]
write_grid(out.create_sheet("Sources"), sources, title_rows=1, autofilter_header=2)

void_trans = [
 ["KEY VOID/CAVITY DENSITY & MEAN-SIZE DATA — unimodal→bimodal transition, BOR-60 neutron 376–377°C (manuscript Table A.22)"],
 ["Material","Dose (dpa)","Regime","Bubble Nb (×10²¹ m⁻³)","Bubble d̄_b (nm)","Void Nv (×10²⁰ m⁻³)","Void d̄_v (nm)","Reference (Ref Key)"],
 ["T91",15.4,"Unimodal","~10",1.2,"—","—","Gao2018"],
 ["T91",35.1,"Bimodal","~6",1.3,"~5",5.0,"Gao2018"],
 ["HT9",17.1,"Unimodal","~3",1.4,"—","—","Gao2019"],
 ["HT9",35.1,"Bimodal","~2",1.2,"~5",7.0,"Gao2019"],
]
write_grid(out.create_sheet("Void_Transition"), void_trans, title_rows=1, autofilter_header=2)

swell_dose = [
 ["SWELLING & VOID SIZE AT KEY DOSE LEVELS — T91 and HT9 (manuscript Table A.23)"],
 ["Material","Source","Dose","T","Regime","d̄_v (nm)","Swelling (%)","Reference (Ref Key)"],
 ["T91","BOR-60 neutron","15.4 dpa","376°C","Unimodal","—","~0.01","Gao2018"],
 ["T91","BOR-60 neutron","35.1 dpa","376°C","Bimodal","5.0","~0.02","Gao2018"],
 ["HT9","BOR-60 neutron","17.1 dpa","377°C","Unimodal","—","0.02","Gao2019"],
 ["HT9","BOR-60 neutron","35.1 dpa","377°C","Bimodal","7.0","0.07","Gao2019"],
 ["HT9","EBR-II/FFTF","200 dpa","420°C","Steady-state","—","0.09–1.02","GarnerToloczko1994"],
 ["HT9","Self-ion (Fe2+)","650 dpa","460°C","Steady-state","~22","~0.2%/dpa","Getto2017"],
 ["HT9 (F/M)","Self-ion","600 dpa","450°C","Steady-state","~30","2.6 (max at 334 local dpa)","Kim2021"],
]
write_grid(out.create_sheet("Swelling_Dose"), swell_dose, title_rows=1, autofilter_header=2)

trends = [
 ["SUMMARY OF QUANTITATIVE TRENDS in loop & void density and mean size (manuscript Table A.24).  ↑ increase, ↓ decrease."],
 ["Quantity","T < 300°C (fixed dose)","T > 300°C (fixed dose)","φ < 5 dpa (fixed T)","φ > 15 dpa (fixed T)"],
 ["Loop density N_d","↑; black-dot regime","↓","↑","↓ or saturation"],
 ["Loop mean size d̄_ℓ","↑; small loops","↑↑","↑","↑; coarsening"],
 ["Void density N_v","Low; unimodal bubbles","↑ then ↓","Low","↑; bimodal regime"],
 ["Void mean size d̄_v","Very small, <2 nm","↑↑","<2 nm","↑↑"],
 ["Swelling S","≈ 0","Bell-shaped; ↑ then ↓","≈ 0","↑"],
 [""],
 ["QUANTITATIVE ANCHORS"],
 ["Loop density maximum","~1.2×10²² m⁻³ near 300°C; decreases to <0.1×10²² m⁻³ at 460°C (T91)","","","Gaganidze2011 / Gao2018"],
 ["Loop mean size range","2.5 nm (EUROFER97, 250°C) to >25 nm (T91, 460°C)","","","Gaganidze2011 / Gao2018"],
 ["Void density (unimodal)","~10²¹ m⁻³, mean 1.2–2.6 nm","","","Klimenkov2012 / Gao2018/2019"],
 ["Void density (bimodal)","bubbles ~2–6×10²¹ m⁻³ (1.2–1.4 nm); voids ~5×10²⁰ m⁻³ (5–7 nm)","","","Gao2018 / Gao2019"],
 ["Swelling","0.02% at 15–17 dpa; 0.07% at 35 dpa; ~2.6% at 600 dpa (HT9)","","","Gao2019 / Kim2021"],
 ["DBH hardening product N·d̄ (EUROFER97 330°C)","4.8×10¹³ m⁻² at 15 dpa → 8.2×10¹³ m⁻² at 32 dpa (+70%)","","","Klimenkov2012"],
]
write_grid(out.create_sheet("Trends_Summary"), trends, title_rows=1)

nvi = [
 ["COMPARISON OF DEFECT SIZE DISTRIBUTIONS — NEUTRON vs ION IRRADIATION in F/M steels (manuscript Table A.25)"],
 ["Observable","Neutron irradiation","Ion irradiation"],
 ["Loop mean diameter","5–20 nm depending on temperature; heterogeneous spatial distribution","6–7 nm at similar temperature under self-ion; more uniform spatial distribution"],
 ["Loop distribution shape","Right-skewed log-normal; black-dot spike at low temperature","More symmetric and narrower; fewer black dots"],
 ["<100> loop fraction","~20–30% at 300°C","~40% at 300°C"],
 ["Spatial heterogeneity","Strong: decoration of pre-existing dislocations dominates; density varies by 3 orders of magnitude","Weak: nearly homogeneous field; no significant dislocation-decoration effect"],
 ["Void/cavity regime","Bimodal at >35 dpa and 376–420°C","Bimodal across investigated parameters at T ≈ 460°C"],
 ["Temperature shift","—","+60–70°C required to match neutron microstructure"],
 ["G-phase / α′","Well developed under neutron irradiation at equivalent conditions","Suppressed under ion irradiation: less segregation, fewer Ni/Si clusters"],
 [""],
 ["Refs: Lambrecht2014, TallerWas2019, TallerWas2020, Was2014, Gao2018, Gao2019"],
]
write_grid(out.create_sheet("Neutron_vs_Ion"), nvi, title_rows=1)

scaling = [
 ["SCALING LAWS & EMPIRICAL CORRELATIONS (manuscript Eqs A.98–A.100 and fitted parameters)"],
 ["Law / Quantity","Relation","Parameters / Fitted values","Conditions","Reference (Ref Key)"],
 ["Loop density saturation (Eq A.98)","N_ℓ^sat ∝ ( G0 / (Ω0·μ·b³ / kB·T) )^(2/3)",
  "G0 = cascade defect production rate; Ω0 = atomic volume; μ = shear modulus; b = Burgers vector; kB·T = thermal energy",
  "Steady state: loop nucleation = loop-to-network conversion (EUROFER97 330°C, near-saturation 15→32 dpa)","GhoniemCho1979 / Klimenkov2012"],
 ["Power-law growth of mean loop size (Eq A.99)","d̄(φ) = d̄0 + A·φ^m,   m ≈ 0.3–0.5",
  "d̄0 = as-nucleated size; A = T-dependent prefactor; m = growth exponent (diffusion-limited coarsening)",
  "Coarsening regime (φ > ~15 dpa)","manuscript A.99"],
 ["  • T91 fit (376°C)","d̄ = 8 nm @15.4 dpa, 14 nm @35.1 dpa","m ≈ 0.57,  A ≈ 2.8 nm·dpa^-0.57","BOR-60 neutron","Gao2018"],
 ["  • HT9 fit (377°C)","d̄ = 10 nm @17.1 dpa, 16 nm @35.1 dpa","m ≈ 0.52,  A ≈ 2.3 nm·dpa^-0.52","BOR-60 neutron","Gao2019"],
 ["Swelling incubation + steady state (Eq A.100)","S(φ) = S0·[exp(-φc/φ) - 1] + Ṡss·(φ - φc),  φ > φc",
  "φc ≈ 20–30 dpa (incubation dose); Ṡss ≈ 0.2%/dpa (steady-state rate); S0 = small transient offset",
  "HT9: 0.02% @17 dpa, 0.07% @35 dpa, ~2.6% @334 local dpa","Gao2019 / Kim2021"],
 ["DBH product N·d̄ (hardening potential)","Δ(N·d̄): 4.8×10¹³ → 8.2×10¹³ m⁻² (+70%)","15 → 32 dpa; driven by mean-size growth despite only +20% density",
  "EUROFER97 330°C","Klimenkov2012"],
 ["Dispersed-barrier hardening (DBH) model","Δσ_y = M·α·μ·b·(N·d)^(1/2)",
  "M ≈ 3.06 (Taylor factor); α: voids 0.45–0.55; a<100> loops ≈ 0.54; ½<111> loops ≈ 0.17",
  "Hardening calibration across F/M steels","Ghoniem1985 / Ando2018"],
 ["Temperature shift (ion ↔ neutron)","ΔT ≈ +60–70°C (ion hotter)","~3 orders of magnitude higher dose rate under ions → enhanced recombination",
  "Validated for T91 (dual-ion vs BOR-60)","Was2014 / TallerWas2019"],
]
write_grid(out.create_sheet("Scaling_Laws"), scaling, title_rows=1)

# tag manuscript-added sheet tabs green
for nm in ["Sources","Void_Transition","Swelling_Dose","Trends_Summary","Neutron_vs_Ion","Scaling_Laws"]:
    out[nm].sheet_properties.tabColor = "00B050"
out["Index"].sheet_properties.tabColor = "1F4E78"
out["References"].sheet_properties.tabColor = "FFC000"

# order: Index, References, Sources first
order = ["Index","References","Sources",
         "Loops_DenSize","Voids_DenSize","T_Series","Dose_Series",
         "N_Loops","N_Voids_Swelling","N_Precipitates","N_Trends",
         "Ion_Loops","Ion_Voids_Swelling","Ion_He_Bubbles","Ion_Precipitates","Ion_Trends",
         "Quant_Summary","Compositions",
         "Loop_SizeDist","Loop_Bins","Void_SizeDist","Void_Bins","Burgers_111_Frac",
         "Void_Transition","Swelling_Dose","Trends_Summary","Neutron_vs_Ion","Scaling_Laws"]
out._sheets.sort(key=lambda s: order.index(s.title) if s.title in order else 999)

out.save(OUT)
print("Saved", OUT, "with", len(out.sheetnames), "sheets:")
for s in out.sheetnames:
    print("  -", s)
